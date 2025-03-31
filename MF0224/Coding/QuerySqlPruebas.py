import mariadb
"""
db = mariadb.connect(
    host="localhost",
    user="root",
    password="0000",
    database="sakila"
)

cursor = db.cursor()
sql = "SELECT c.customer_id, c.first_name, c.last_name, COUNT(r.rental_id) AS Rents FROM customer c JOIN rental r ON c.customer_id = r.customer_id GROUP BY c.customer_id ORDER BY Rents DESC"
cursor.execute(sql)
result = cursor.fetchall()
for row in result:
    print(row)

db = mariadb.connect(
    host="localhost",
    user="root",
    password="0000",
    database="information_schema"
)

cursor = db.cursor()


sql = "SELECT t.TABLE_NAME, t.TABLE_TYPE, c.`COLUMN_NAME`, c.DATA_TYPE FROM `TABLES` t JOIN COLUMNS c ON c.`TABLE_NAME`= t.`TABLE_NAME` WHERE t.table_schema = 'sakila'"
cursor.execute(sql)
result = cursor.fetchall()
for row in result:
    print(row)


db = mariadb.connect(
    host="localhost",
    user="root",
    password="0000",
    database="information_schema"
)

db_info = mariadb.connect(
    host="localhost",
    user="root",
    password="0000",
    database="information_schema"
)

cursor = db_info.cursor()

sql = "SELECT t.TABLE_NAME, t.TABLE_TYPE, c.`COLUMN_NAME`, c.DATA_TYPE FROM `TABLES` t JOIN COLUMNS c ON c.`TABLE_NAME`= t.`TABLE_NAME` WHERE t.table_schema = 'sakila'"
cursor.execute(sql)
result = cursor.fetchall()
for row in result:
    print(row)"
    """
""" ## Funzionante con gi´tutte le Query
# Connect to the database information_schema
db = mariadb.connect(
    host="localhost",
    user="root",
    password="0000",
    database="information_schema"
)

# Function to execute a query and return the results
def execute_query(connection, query):
        cursor = connection.cursor()
        cursor.execute(query)
        return cursor.fetchall()

# Open a file to write the results

# First Query
q1 = "SELECT * FROM `TABLES` t WHERE table_schema = 'sakila';"
r1 = execute_query(db, q1)
if r1:
    print("Resultados primera Query:")
    for row in r1:
        print(row)

# Second Query
q2 = "SELECT t.TABLE_NAME, t.TABLE_TYPE, c.`COLUMN_NAME`, c.DATA_TYPE FROM `TABLES` t " \
"JOIN COLUMNS c ON c.`TABLE_NAME`= t.`TABLE_NAME` WHERE t.table_schema = 'sakila';"
r2 = execute_query(db, q2)
if r2:
    print("\nResultados segunda Query:")
    for row2 in r2:
        print(row2)

# Third Query
q3 = "SELECT k.TABLE_SCHEMA, k.`TABLE_NAME`, k.`COLUMN_NAME`, k.`CONSTRAINT_NAME` " \
"FROM KEY_COLUMN_USAGE k WHERE k.TABLE_SCHEMA = 'sakila'ORDER BY k.`CONSTRAINT_NAME` DESC;"
r3 = execute_query(db, q3)
if r3:
    print("\nResultados tercera Query:")
    for row3 in r3:
        print(row3)

# Fourth Query
q4 = "SELECT k.`TABLE_NAME`, k.`COLUMN_NAME`, count(k.`CONSTRAINT_NAME`) AS Key_Num " \
"FROM KEY_COLUMN_USAGE k WHERE k.TABLE_SCHEMA = 'sakila' GROUP BY k.`CONSTRAINT_NAME` ORDER BY k.`CONSTRAINT_NAME` DESC;"
r4 = execute_query(db, q4)
if r4:
    print("\nResultados cuarta Query:")
    for row4 in r4:
        print(row4)
"""
"""## stampante u .txt
import mariadb

# Connect to the database information_schema
db = mariadb.connect(
    host="localhost",
    user="root",
    password="0000",
    database="information_schema"
)

# Function to execute a query and return the results
def execute_query(connection, query):
    cursor = connection.cursor()
    cursor.execute(query)
    return cursor.fetchall()

# Open a file to write the results
with open("query_results.txt", "w", encoding="utf-8") as file:
    # First Query
    q1 = "SELECT * FROM `TABLES` t WHERE table_schema = 'sakila';"
    r1 = execute_query(db, q1)
    if r1:
        file.write("Resultados primera Query:\n")
        for row in r1:
            file.write(f"{row}\n")

    # Second Query
    q2 = "SELECT t.TABLE_NAME, t.TABLE_TYPE, c.`COLUMN_NAME`, c.DATA_TYPE FROM `TABLES` t " \
         "JOIN COLUMNS c ON c.`TABLE_NAME`= t.`TABLE_NAME` WHERE t.table_schema = 'sakila';"
    r2 = execute_query(db, q2)
    if r2:
        file.write("\nResultados segunda Query:\n")
        for row2 in r2:
            file.write(f"{row2}\n")

    # Third Query
    q3 = "SELECT k.TABLE_SCHEMA, k.`TABLE_NAME`, k.`COLUMN_NAME`, k.`CONSTRAINT_NAME` " \
         "FROM KEY_COLUMN_USAGE k WHERE k.TABLE_SCHEMA = 'sakila' ORDER BY k.`CONSTRAINT_NAME` DESC;"
    r3 = execute_query(db, q3)
    if r3:
        file.write("\nResultados tercera Query:\n")
        for row3 in r3:
            file.write(f"{row3}\n")

    # Fourth Query
    q4 = "SELECT k.`TABLE_NAME`, k.`COLUMN_NAME`, count(k.`CONSTRAINT_NAME`) AS Key_Num " \
         "FROM KEY_COLUMN_USAGE k WHERE k.TABLE_SCHEMA = 'sakila' GROUP BY k.`CONSTRAINT_NAME` ORDER BY k.`CONSTRAINT_NAME` DESC;"
    r4 = execute_query(db, q4)
    if r4:
        file.write("\nResultados cuarta Query:\n")
        for row4 in r4:
            file.write(f"{row4}\n")

print("I risultati sono stati salvati in 'query_results.txt'.")
"""
"""## stampante u .xlsx
from openpyxl import Workbook

# Connect to the database information_schema
db = mariadb.connect(
    host="localhost",
    user="root",
    password="0000",
    database="information_schema"
)

# Function to execute a query and return the results
def execute_query(connection, query):
    cursor = connection.cursor()
    cursor.execute(query)
    return cursor.fetchall()

# Create a new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Query Results"

# First Query
q1 = "SELECT c.`TABLE_NAME`, c.`COLUMN_NAME`, c.DATA_TYPE, c.CHARACTER_SET_NAME, c.COLUMN_TYPE, c.COLUMN_KEY, c.EXTRA, c.`PRIVILEGES` " \
"FROM information_schema.COLUMNS c " \
"WHERE TABLE_SCHEMA = 'sakila';"
r1 = execute_query(db, q1)
if r1:
    ws.append(["Resultados primera Query:"]),  # Add a header
    for row in r1:
        ws.append(row)  # Add each row to the Excel sheet

# Second Query
q2 = "SELECT t.TABLE_NAME, t.TABLE_TYPE, c.`COLUMN_NAME`, c.DATA_TYPE FROM `TABLES` t " \
     "JOIN COLUMNS c ON c.`TABLE_NAME`= t.`TABLE_NAME` WHERE t.table_schema = 'sakila';"
r2 = execute_query(db, q2)
if r2:
    ws.append([])  # Add an empty row for separation
    ws.append(["Resultados segunda Query:"])  # Add a header
    for row2 in r2:
        ws.append(row2)

# Third Query
q3 = "SELECT k.TABLE_SCHEMA, k.`TABLE_NAME`, k.`COLUMN_NAME`, k.`CONSTRAINT_NAME` " \
     "FROM KEY_COLUMN_USAGE k WHERE k.TABLE_SCHEMA = 'sakila' ORDER BY k.`CONSTRAINT_NAME` DESC;"
r3 = execute_query(db, q3)
if r3:
    ws.append([])  # Add an empty row for separation
    ws.append(["Resultados tercera Query:"])  # Add a header
    for row3 in r3:
        ws.append(row3)

# Fourth Query
q4 = "SELECT k.`TABLE_NAME`, k.`COLUMN_NAME`, count(k.`CONSTRAINT_NAME`) AS Key_Num " \
     "FROM KEY_COLUMN_USAGE k WHERE k.TABLE_SCHEMA = 'sakila' GROUP BY k.`CONSTRAINT_NAME` ORDER BY k.`CONSTRAINT_NAME` DESC;"
r4 = execute_query(db, q4)
if r4:
    ws.append([])  # Add an empty row for separation
    ws.append(["Resultados cuarta Query:"])  # Add a header
    for row4 in r4:
        ws.append(row4)

# Save the workbook to a file
wb.save("query_results.xlsx")
print("I risultati sono stati salvati in 'query_results.xlsx'.")
"""

from openpyxl import Workbook
import mariadb

# Connect to the database information_schema
db = mariadb.connect(
    host="localhost",
    user="root",
    password="0000",
    database="information_schema"
)

# Function to execute a query and return the results
def execute_query(connection, query):
    cursor = connection.cursor()
    cursor.execute(query)
    return cursor.fetchall(), [desc[0] for desc in cursor.description]  # Return data and column names

# Create a new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Query Results"

# First Query
q1 = "SELECT c.`TABLE_NAME`, c.`COLUMN_NAME`, c.DATA_TYPE, c.CHARACTER_SET_NAME, c.COLUMN_TYPE, c.COLUMN_KEY, c.EXTRA, c.`PRIVILEGES` " \
     "FROM information_schema.COLUMNS c " \
     "WHERE TABLE_SCHEMA = 'sakila';"
r1, columns1 = execute_query(db, q1)
if r1:
    ws.append(["Todas las Tablas:"])  # Add a header
    ws.append(columns1)  # Add column titles
    for row in r1:
        ws.append(row)  # Add each row to the Excel sheet

# Second Query
q2 = "SELECT t.TABLE_NAME, t.TABLE_TYPE, c.`COLUMN_NAME`, c.DATA_TYPE FROM `TABLES` t " \
     "JOIN COLUMNS c ON c.`TABLE_NAME`= t.`TABLE_NAME` " \
     "WHERE t.table_schema = 'sakila';"
r2, columns2 = execute_query(db, q2)
if r2:
    ws.append([])  # Add an empty row for separation
    ws.append(["Resultados segunda Query:"])  # Add a header
    ws.append(columns2)  # Add column titles
    for row2 in r2:
        ws.append(row2)

# Third Query
q3 = "SELECT k.TABLE_SCHEMA, k.`TABLE_NAME`, k.`COLUMN_NAME`, k.`CONSTRAINT_NAME` " \
     "FROM KEY_COLUMN_USAGE k " \
     "WHERE k.TABLE_SCHEMA = 'sakila' " \
     "ORDER BY k.`CONSTRAINT_NAME` DESC;"
r3, columns3 = execute_query(db, q3)
if r3:
    ws.append([])  # Add an empty row for separation
    ws.append(["Resultados tercera Query:"])  # Add a header
    ws.append(columns3)  # Add column titles
    for row3 in r3:
        ws.append(row3)

# Fourth Query
q4 = "SELECT k.`TABLE_NAME`, k.`COLUMN_NAME`, count(k.`CONSTRAINT_NAME`) AS Key_Num " \
     "FROM KEY_COLUMN_USAGE k WHERE k.TABLE_SCHEMA = 'sakila' GROUP BY k.`CONSTRAINT_NAME` ORDER BY k.`CONSTRAINT_NAME` DESC;"
r4, columns4 = execute_query(db, q4)
if r4:
    ws.append([])  # Add an empty row for separation
    ws.append(["Resultados cuarta Query:"])  # Add a header
    ws.append(columns4)  # Add column titles
    for row4 in r4:
        ws.append(row4)

# Save the workbook to a file
wb.save("query_results.xlsx")
print("I risultati sono stati salvati in 'query_results.xlsx'.")

""" excel con titolo
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import mariadb

# Connect to the database information_schema
db = mariadb.connect(
    host="localhost",
    user="root",
    password="0000",
    database="information_schema"
)

# Function to execute a query and return the results
def execute_query(connection, query):
    cursor = connection.cursor()
    cursor.execute(query)
    return cursor.fetchall(), [desc[0] for desc in cursor.description]  # Return data and column names

# Create a new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Query Results"

# Function to add a title for each query
def add_query_title(sheet, title, start_col, end_col, row):
    sheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = sheet.cell(row=row, column=start_col)
    cell.value = title
    cell.font = Font(size=12, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# First Query
q1 = "SELECT c.`TABLE_NAME`, c.`COLUMN_NAME`, c.DATA_TYPE, c.CHARACTER_SET_NAME, c.COLUMN_TYPE, c.COLUMN_KEY, c.EXTRA, c.`PRIVILEGES` " \
     "FROM information_schema.COLUMNS c " \
     "WHERE TABLE_SCHEMA = 'sakila';"
r1, columns1 = execute_query(db, q1)
if r1:
    add_query_title(ws, "Resultados primera Query", 1, len(columns1), ws.max_row + 1)  # Add title
    ws.append(columns1)  # Add column titles
    for row in r1:
        ws.append(row)  # Add each row to the Excel sheet

# Second Query
q2 = "SELECT t.TABLE_NAME, t.TABLE_TYPE, c.`COLUMN_NAME`, c.DATA_TYPE FROM `TABLES` t " \
     "JOIN COLUMNS c ON c.`TABLE_NAME`= t.`TABLE_NAME` " \
     "WHERE t.table_schema = 'sakila';"
r2, columns2 = execute_query(db, q2)
if r2:
    ws.append([])  # Add an empty row for separation
    add_query_title(ws, "Resultados segunda Query", 1, len(columns2), ws.max_row + 1)  # Add title
    ws.append(columns2)  # Add column titles
    for row2 in r2:
        ws.append(row2)

# Third Query
q3 = "SELECT k.TABLE_SCHEMA, k.`TABLE_NAME`, k.`COLUMN_NAME`, k.`CONSTRAINT_NAME` " \
     "FROM KEY_COLUMN_USAGE k " \
     "WHERE k.TABLE_SCHEMA = 'sakila' " \
     "ORDER BY k.`CONSTRAINT_NAME` DESC;"
r3, columns3 = execute_query(db, q3)
if r3:
    ws.append([])  # Add an empty row for separation
    add_query_title(ws, "Resultados tercera Query", 1, len(columns3), ws.max_row + 1)  # Add title
    ws.append(columns3)  # Add column titles
    for row3 in r3:
        ws.append(row3)

# Fourth Query
q4 = "SELECT k.`TABLE_NAME`, k.`COLUMN_NAME`, count(k.`CONSTRAINT_NAME`) AS Key_Num " \
     "FROM KEY_COLUMN_USAGE k WHERE k.TABLE_SCHEMA = 'sakila' GROUP BY k.`CONSTRAINT_NAME` ORDER BY k.`CONSTRAINT_NAME` DESC;"
r4, columns4 = execute_query(db, q4)
if r4:
    ws.append([])  # Add an empty row for separation
    add_query_title(ws, "Resultados cuarta Query", 1, len(columns4), ws.max_row + 1)  # Add title
    ws.append(columns4)  # Add column titles
    for row4 in r4:
        ws.append(row4)

# Save the workbook to a file
wb.save("query_results.xlsx")
print("I risultati sono stati salvati in 'query_results.xlsx'.")"
"""